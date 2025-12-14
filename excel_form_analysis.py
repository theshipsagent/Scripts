import openpyxl
import re
import pandas as pd
from pathlib import Path
import sys
import tkinter as tk
from tkinter import filedialog
import logging
from datetime import datetime
import os

try:
    import xlrd  # Optional for .xls files
except ImportError:
    xlrd = None


def setup_logging():
    """Configure logging to save to specified directory with timestamp."""
    log_dir = Path(r"C:\Users\wsd3\OneDrive\GRoK\Logs")
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"formula_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logging.basicConfig(
        filename=log_file,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    logging.info("Logging initialized.")


def select_file():
    """Open file explorer to select Excel file."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xls *.xlsx *.xlsm")]
    )
    root.destroy()
    if not file_path:
        logging.error("No file selected. Exiting.")
        print("No file selected. Exiting.")
        sys.exit(1)
    logging.info(f"Selected file: {file_path}")
    return file_path


def extract_formula_details(formula):
    """Extract functions and cell references from an Excel formula."""
    try:
        # Common Excel time-related functions
        time_funcs = {'NOW', 'TODAY', 'TIME', 'HOUR', 'MINUTE', 'SECOND', 'DATEDIF', 'NETWORKDAYS', 'WORKDAY'}
        # Find all functions (e.g., SUM, TIME) in formula
        functions = re.findall(r'[A-Z]+(?=\()', formula)
        # Find cell references (e.g., A1, B2:C5)
        cell_refs = re.findall(r'[A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?', formula)
        # Check if formula involves time-related functions
        is_time_related = any(func in time_funcs for func in functions)
        return {
            'functions': functions,
            'cell_references': cell_refs,
            'is_time_related': is_time_related
        }
    except Exception as e:
        logging.error(f"Error extracting formula details: {str(e)}")
        return {'functions': [], 'cell_references': [], 'is_time_related': False, 'error': str(e)}


def analyze_excel_formulas():
    """Analyze all formulas in selected Excel file and save results to CSV."""
    setup_logging()
    try:
        file_path = select_file()
        logging.info(f"Starting analysis for file: {file_path}")

        # Check file extension
        ext = Path(file_path).suffix.lower()
        if ext == '.xls' and xlrd:
            logging.warning(
                "Detected .xls file. Attempting xlrd fallback. Consider converting to .xlsx for better compatibility.")
            print("Warning: .xls file detected. Consider converting to .xlsx for better compatibility.")
            try:
                wb = xlrd.open_workbook(file_path)
                logging.info("Loaded .xls file with xlrd. Limited formula support.")
                print("Note: .xls formula extraction is limited. Convert to .xlsx for full analysis.")
            except Exception as e:
                logging.error(f"Failed to load .xls with xlrd: {str(e)}")
                print(f"Error: Failed to load .xls. Please convert to .xlsx and retry. {str(e)}")
                sys.exit(1)
        else:
            # Try openpyxl for .xlsx/.xlsm or .xls
            try:
                wb = openpyxl.load_workbook(file_path, data_only=False)
                logging.info("Loaded workbook with openpyxl.")
            except Exception as e:
                logging.error(f"Failed to load with openpyxl: {str(e)}")
                print(f"Error: Failed to load file. If .xls, convert to .xlsx or install xlrd. {str(e)}")
                sys.exit(1)

        results = []
        # Handle openpyxl workbook
        if isinstance(wb, openpyxl.Workbook):
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                logging.info(f"Processing sheet: {sheet_name}")
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Check if cell contains a formula
                            formula = cell.value
                            details = extract_formula_details(formula)
                            results.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'formula': formula,
                                'functions': ', '.join(details['functions']),
                                'cell_references': ', '.join(details['cell_references']),
                                'is_time_related': details['is_time_related'],
                                'error': details.get('error', '')
                            })
        else:
            # Handle xlrd workbook (limited formula support)
            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                logging.info(f"Processing sheet: {sheet_name}")
                print(f"Warning: xlrd does not support direct formula extraction. Results may be incomplete.")
                logging.warning("xlrd does not support direct formula extraction.")
                results.append({
                    'sheet': sheet_name,
                    'cell': 'N/A',
                    'formula': 'N/A',
                    'functions': '',
                    'cell_references': '',
                    'is_time_related': False,
                    'error': 'xlrd does not support formula extraction'
                })

        # Save results to DataFrame and CSV
        df = pd.DataFrame(results)
        output_path = Path(file_path).parent / f"{Path(file_path).stem}_formula_analysis.csv"
        df.to_csv(output_path, index=False)
        logging.info(f"Analysis saved to {output_path}")
        print(f"Analysis saved to {output_path}")

        # Summarize time-related formulas
        time_formulas = df[df['is_time_related'] == True]
        if not time_formulas.empty:
            print("\nTime-related formulas found:")
            logging.info("Time-related formulas found:")
            for _, row in time_formulas.iterrows():
                print(f"Sheet: {row['sheet']}, Cell: {row['cell']}, Formula: {row['formula']}")
                logging.info(f"Sheet: {row['sheet']}, Cell: {row['cell']}, Formula: {row['formula']}")
        else:
            print("\nNo time-related formulas detected.")
            logging.info("No time-related formulas detected.")

        return df

    except FileNotFoundError:
        logging.error(f"File '{file_path}' not found.")
        print(f"Error: File '{file_path}' not found.")
        sys.exit(1)
    except Exception as e:
        logging.error(f"Error: {str(e)}")
        print(f"Error: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    analyze_excel_formulas()